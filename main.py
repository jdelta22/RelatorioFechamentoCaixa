import tabula
import pandas as pd
import openpyxl
from openpyxl.styles import NamedStyle, PatternFill, Alignment, Font, Border, Side
import pdfplumber
import re
from datetime import datetime
import jpype



PDF_FILE = "VendasPorProdutos.pdf"
EXCEL_MODELO = "CAIXA_format.xlsx"
EXCEL_SAIDA = "CAIXA_fechado.xlsx"
PDF_RECEBIMENTOS = "RelatorioFechamentoCaixaBobina.pdf"



def ler_pdf_produto(caminho_pdf):
    """Lê o PDF e retorna o DataFrame tratado."""
    try:
        tabelas = tabula.read_pdf(caminho_pdf, pages="all")
        print("✅ Arquivo PDF lido com sucesso!")
    except Exception as e:
        print(f"❌ Erro ao ler o PDF: {caminho_pdf},{e}")
        return None

    tabela = tabelas[0]
    colunas_remover = ["Barras", "Custo", "% Média", "Lucro bruto"]
    tabela = tabela.drop(columns=colunas_remover, errors='ignore')

    # Extrai código
    tabela[['Codigo', '_']] = tabela["Código UN"].str.split(" ", expand=True)
    tabela['Codigo'] = tabela['Codigo'].astype(str).str[-6:]
    tabela = tabela.drop(columns=['Código UN', '_'], errors='ignore')
    tabela = tabela.set_index("Codigo")

    # Limpeza numérica
    tabela['Qtd'] = tabela['Qtd'].astype(str).str.replace(',000', '', regex=False)
    tabela['Qtd'] = pd.to_numeric(tabela['Qtd'], errors='coerce')
    tabela['Venda'] = tabela['Venda'].astype(str).str.replace(',', '.')
    tabela['Venda'] = pd.to_numeric(tabela['Venda'], errors='coerce')

    return tabela

def ler_pdf_recebimentos(caminho_pdf):
    try:
        with pdfplumber.open("RelatorioFechamentoCaixaBobina.pdf") as pdf:
            page = pdf.pages[0]
            recebimentos = page.extract_text()
            print("✅ Arquivo PDF lido com sucesso!")
    except Exception as e:
        print(f"❌ Erro ao ler o PDF: {caminho_pdf},{e}")
        return None
    
    return recebimentos

def configurar_estilos(caixa):
    """Aplica estilos nas células do Excel."""
    moeda_style = NamedStyle(name='moeda', number_format='"R$" #,##0.00;[Red]"R$" -#,##0.00')
    moeda_style.alignment = Alignment(horizontal='right', vertical='bottom')
    moeda_style.font = Font(name='Calibri', size=14)
    fundo = PatternFill(fill_type="solid", fgColor="ffff00")

    def aplicar_estilo(faixa, cor=False, borda_tipo="fina"):
        """
        Aplica estilo de moeda, cor de fundo opcional e borda personalizada
        nas células do intervalo informado.
        
        faixa: ex. "C4:C41" ou "C42"
        cor: se True, aplica cor de fundo
        borda_tipo: "fina", "grossa" ou None
        """
        intervalo = caixa[faixa]

        # 🔧 Ajusta se for uma única célula
        if not isinstance(intervalo, tuple):
            intervalo = [[intervalo]]

        # 🧱 Define bordas conforme tipo
        if borda_tipo == "grossa":
            borda = Border(
                left=Side(style="medium"),
                right=Side(style="medium"),
                top=Side(style="medium"),
                bottom=Side(style="medium")
            )
        elif borda_tipo == "fina":
            borda = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        else:
            borda = None

        # 🖌️ Aplica os estilos
        for row in intervalo:
            for cell in row:
                cell.style = moeda_style
                if cor:
                    cell.fill = fundo
                if borda:
                    cell.border = borda

    aplicar_estilo('C4:C45') # valores de unidades
    aplicar_estilo('G4:G10') # valores de caixas
    aplicar_estilo('B51:B56') #valores de despesas
    aplicar_estilo('B60:B63') # valores de recebimentos
    aplicar_estilo('F11:G11', cor=True, borda_tipo ='grossa') # valor total de caixas
    aplicar_estilo('C46', cor=True, borda_tipo ='grossa') # valor total de unidades
    aplicar_estilo('C48', cor=True, borda_tipo ='grossa') # valor total de caixas + unidades
    aplicar_estilo('B57', cor=True, borda_tipo ='grossa') # valor total de despesas
    aplicar_estilo('B64', cor=True, borda_tipo ='grossa') # valor total de recebimentos
    aplicar_estilo('C66', cor=True, borda_tipo ='grossa') # valor total de recebimentos em dinheiro - despesas

def acertar_data(caixa):
    data_atual = datetime.now().strftime("%d/%m/%Y")
    caixa.merge_cells('A1:G1')
    caixa['A1'] = f"CAIXA DIA {data_atual}"
    caixa['A1'].font = Font(bold=True, size=16)
    caixa['A1'].alignment = Alignment(horizontal='center', vertical='center')
    caixa['A1'].border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"))

def preencher_dados_produtos(caixa, tabela):
    """Preenche os valores de quantidade e venda com base nos códigos."""
    mapeamento = {
        '000001': 'B4', # doce
        '000002': 'B5', # sal
        '000003': 'B6', # suíça
        '000007': 'B7', # nazare
        '000011': 'B8', # queijo
        '000012': 'B9', # coquinho
        '000026': 'B10', # sequilho
        '000063': 'B11', # fardo
        '000027': 'B12', # recheado
        '000052': 'B13', # rosquinha
        '000028': 'B14', # sua mae
        '000086': 'B15', # bolo grande
        '000037': 'B16', # bolo pequeno
        '000114': 'B17', # doce de leite pequeno
        '000040': 'B18', # doce de leite pote
        '000057': 'B19', # amendoin
        '000033': 'B20', # castanha
        '000085': 'B21', # nego bom
        '000122': 'B22', # pé de moça
        '000123': 'B23', # beiju recheado
        '000117': 'B24', # biscoito alexandre
        '000124': 'B25', # brigadeiro
        '000121': 'B26', # pingo bel
        '000068': 'B27', # frigelis preto
        '000070': 'B28', # frigelis verde
        '000069': 'B29', # frigelis vermelho
        '000118': 'B30', # cocada grande
        '000112': 'B31', # cocada pequena
        '000073': 'B32', # bala de yogurte
        '000087': 'B33', # pirulito
        '000126': 'B34', # coqueiro
        '000127': 'B35', # zambanana
        '000075': 'B36', # coca
        '000076': 'B37', # coca zero
        '000079': 'B38', # fanta
        '000078': 'B39', # sukita
        '000077': 'B40', # guaraná
        '000093': 'B41', # limoneto
        '000080': 'B42', # agua
        '000082': 'B43', # agua c/ gás
        '000095': 'B44', # caixa de presente
        '000096': 'B45', # estojo bolacho
        # CAIXAS
        '000009': 'F4', # Nazaré
        '000010': 'F5', # Queijo
        '000031': 'F6', # Coquinho
        '000004': 'F7', # Doce
        '000005': 'F8', # Sal
        '000006': 'F9', # Suíça
        '000025': 'F10', # Mista
    }

    for codigo, celula in mapeamento.items():
        try:
            coluna_qtd = celula[0]
            linha = celula[1:]
            if coluna_qtd == 'B':
                caixa[f'B{linha}'] = tabela['Qtd'][codigo]
                caixa[f'C{linha}'] = tabela['Venda'][codigo]
            elif coluna_qtd == 'F':
                caixa[f'F{linha}'] = tabela['Qtd'][codigo]
                caixa[f'G{linha}'] = tabela['Venda'][codigo]
        except KeyError:
            continue

def organizar_dados_recebimentos(recebimentos):
    def parse_dav(recebimentos):
        padrao = r"[A-Z ]+\s+([A-Z]{2})\s+(\d{1,3}(?:\.\d{3})*,\d{2})"
        return re.findall(padrao, recebimentos)

    def parse_sangrias(recebimentos):
        padrao = r"DAV\s+([A-Z ]+)\s+(\d{1,3}(?:\.\d{3})*,\d{2})"
        return re.findall(padrao, recebimentos)

    def separar_secoes(recebimentos):
        # Marcações claras no relatório
        secoes = {
            "dav": "",
            "sangrias": "",
            "totalizadores": ""
        }

        # Quebra em blocos
        linhas = recebimentos.splitlines()

        secao_atual = None

        for linha in linhas:
            l = linha.strip().upper()

            if "DAV" in l and "VENDAS" in l:
                secao_atual = "dav"
                continue

            if "LISTA DAS SANGRIAS" in l:
                secao_atual = "sangrias"
                continue

            if "TOTALIZADORES" in l:
                secao_atual = "totalizadores"
                continue

            if secao_atual:
                secoes[secao_atual] += linha + "\n"

        return secoes

    def parse_relatorio(recebimentos):
        secoes = separar_secoes(recebimentos)

        # DAV (primeiro bloco)
        dav = parse_dav(secoes["dav"])

        # Sangrias
        sangrias = parse_sangrias(secoes["sangrias"])

        # Totalizadores (parte final)
        totalizadores = parse_dav(secoes["totalizadores"])

        return {
            "dav": dav,
            "sangrias": sangrias,
            "totalizadores": totalizadores
        }
    
    dav = dict(parse_relatorio(recebimentos)['dav'])
    sangria = dict(parse_relatorio(recebimentos)['sangrias'])

    return dav, sangria
    
def preencher_dados_recebimentos( caixa, dav, sangria):
    for ab, val in dav.items():
        if ab == 'DN':
            caixa['B60'] = val
        elif ab == 'CD':
            caixa['B61'] = val
        elif ab == 'PX':
            caixa['B62'] = val
        elif ab == 'PZ':
            caixa['B63'] = val

    linha_init = 51
    for i, (desc, valor) in enumerate(sangria.items()):
        linha = linha_init + i
        if linha > 56:
            break  # evita ultrapassar as 6 linhas da planilha

        caixa[f"A{linha}"] = desc.strip()
        caixa[f"B{linha}"] = valor
    
def gerar_excel(tabela, recebimentos):
    """Carrega modelo, aplica dados e salva resultado."""
    try:
        relatorio = openpyxl.load_workbook(EXCEL_MODELO, data_only=False)
        caixa = relatorio['Planilha1']
        configurar_estilos(caixa)
        acertar_data(caixa)
        preencher_dados_produtos(caixa, tabela)
        dav, sangria = organizar_dados_recebimentos(recebimentos)
        preencher_dados_recebimentos(caixa, dav, sangria)
        relatorio.save(EXCEL_SAIDA)
        print(f"✅ Arquivo gerado com sucesso: {EXCEL_SAIDA}")
    except Exception as e:
        print(f"❌ Erro ao gerar o Excel: {e}")


def main():
    print(f"Lendo o arquivo `{PDF_FILE}`...\n")
    tabela = ler_pdf_produto(PDF_FILE)
    print(f"Lendo o arquivo `{PDF_RECEBIMENTOS}`...\n")
    recebimentos = ler_pdf_recebimentos(PDF_RECEBIMENTOS)
    if tabela is not None and recebimentos is not None:
        gerar_excel(tabela, recebimentos)
    input("\nPressione ENTER para fechar...")


if __name__ == "__main__":
    main()